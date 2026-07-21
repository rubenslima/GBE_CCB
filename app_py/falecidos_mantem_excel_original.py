"""
Projeto         : Comunicado falecimento
Autor           : Rubens Lima
Solicitante     : Silvia
Criado em       : 2026-06-10
Última alteração: 2026-06-10
Versão          : 1.1.0
Descrição       : Recebe uma planilha Excel, preserva suas abas e adiciona DATAINCLUSAO e DATADEFERIMENTO
Tipo            : ETL
Módulo          : Conferência CCB
Tags            : falecidos
ID              : GBE.CCB.20260610.001
"""

import os
import re
import time
import urllib.parse
from pathlib import Path

import pandas as pd
from dotenv import load_dotenv
from sqlalchemy import create_engine, text
from sqlalchemy.engine import Engine
import openpyxl


# ------------------------------
# Utilidades
# ------------------------------
def limpar_console() -> None:
    os.system("cls" if os.name == "nt" else "clear")


nome_arquivo = os.path.splitext(os.path.basename(__file__))[0]


def carregar_cfg():
    load_dotenv()
    cfg = {
        "SERVER": (os.getenv("SERVER") or "").strip(),
        "USER": (os.getenv("USER") or "").strip(),
        "PASSWORD": (os.getenv("PASSWORD") or "").strip(),
        "DATABASE": (os.getenv("DATABASE") or "").strip(),
        "ODBC_DRIVER": (
            os.getenv("ODBC_DRIVER") or "ODBC Driver 17 for SQL Server"
        ).strip(),
        "ODBC_EXTRA": (
            os.getenv("ODBC_EXTRA") or ""
        ).strip(),  # ex.: Encrypt=yes;TrustServerCertificate=yes
    }
    faltando = [k for k in ("SERVER", "USER", "PASSWORD", "DATABASE") if not cfg[k]]
    if faltando:
        raise RuntimeError(f"Variáveis ausentes no .env: {', '.join(faltando)}")
    return cfg


def build_connection_url(cfg) -> str:
    """
    Usa DSN-less com odbc_connect e quote_plus para evitar problemas com caracteres especiais.
    """
    params = (
        f"DRIVER={{{cfg['ODBC_DRIVER']}}};"
        f"SERVER={cfg['SERVER']};"
        f"DATABASE={cfg['DATABASE']};"
        f"UID={cfg['USER']};"
        f"PWD={cfg['PASSWORD']}"
    )
    if cfg["ODBC_EXTRA"]:
        extra = cfg["ODBC_EXTRA"]
        if not extra.endswith(";"):
            extra += ";"
        params += ";" + extra
    return f"mssql+pyodbc:///?odbc_connect={urllib.parse.quote_plus(params)}"


def get_engine(cfg) -> Engine:
    url = build_connection_url(cfg)
    return create_engine(
        url,
        pool_pre_ping=True,
        pool_recycle=1800,
        pool_size=5,
        max_overflow=5,
    )


def garantir_pasta(caminho: str) -> None:
    os.makedirs(caminho, exist_ok=True)


def sanitize_filename(name: str) -> str:
    # Remove caracteres inválidos em nomes de arquivo (Windows-safe)
    return re.sub(r'[<>:"/\\|?*]+', "-", name).strip() or "export"


def limpar_nome_coluna(coluna: str) -> str:
    """
    Normaliza nome de coluna para comparação:
    - remove espaços
    - remove acentos simples
    - deixa em maiúsculo
    """
    texto = str(coluna).strip().upper()
    substituicoes = {
        "Á": "A",
        "À": "A",
        "Â": "A",
        "Ã": "A",
        "É": "E",
        "Ê": "E",
        "Í": "I",
        "Ó": "O",
        "Ô": "O",
        "Õ": "O",
        "Ú": "U",
        "Ç": "C",
    }
    for origem, destino in substituicoes.items():
        texto = texto.replace(origem, destino)
    texto = re.sub(r"[^A-Z0-9]", "", texto)
    return texto


def localizar_coluna_matricula(df: pd.DataFrame) -> str | None:
    """
    Localiza uma coluna de matrícula, aceitando nomes como:
    MATRICULA, Matrícula, NUM_MATRICULA, CD_MATRICULA.
    """
    candidatos = {
        "MATRICULA",
        "NUMMATRICULA",
        "CDMATRICULA",
        "NUMEROMATRICULA",
    }

    for coluna in df.columns:
        if limpar_nome_coluna(coluna) in candidatos:
            return coluna

    return None


def normalizar_matricula(valor) -> str:
    """
    Mantém apenas dígitos e preenche com zeros à esquerda até 9 posições.

    """
    if pd.isna(valor):
        return ""

    texto = str(valor).strip()

    # Trata casos em que o Excel leia como número: 12345.0
    if re.fullmatch(r"\d+\.0", texto):
        texto = texto[:-2]

    somente_digitos = re.sub(r"\D", "", texto)

    if not somente_digitos:
        return ""

    return somente_digitos.zfill(9)[-9:]


def autosize_columns(writer, sheet_name: str, df: pd.DataFrame) -> None:
    """
    Ajuste simples de largura por coluna usando xlsxwriter.
    """
    ws = writer.sheets[sheet_name]
    for idx, col in enumerate(df.columns, start=0):
        if df.empty:
            max_len = len(str(col)) + 2
        else:
            serie = df[col].astype(str)
            max_len = max(serie.map(len).max(), len(str(col))) + 2
        ws.set_column(idx, idx, min(max_len, 60))


def nome_aba_excel_valido(nome: str) -> str:
    """
    Garante nome de aba válido para Excel.
    """
    nome_limpo = re.sub(r"[\\/*?:\[\]]", "", str(nome)).strip()
    return (nome_limpo or "Aba")[:31]


def solicitar_arquivo_excel() -> Path:
    caminho = (
        input("Informe o caminho completo do arquivo Excel de entrada: ")
        .strip()
        .strip('"')
    )

    if not caminho:
        raise RuntimeError("Nenhum arquivo informado.")

    arquivo = Path(caminho)

    if not arquivo.exists():
        raise RuntimeError(f"Arquivo não encontrado: {arquivo}")

    if arquivo.suffix.lower() not in [".xlsx", ".xlsm", ".xls"]:
        raise RuntimeError(
            "O arquivo informado precisa ser Excel: .xlsx, .xlsm ou .xls"
        )

    return arquivo


def buscar_datas_comunicado_falecimento(engine: Engine) -> pd.DataFrame:
    query = """
        SET DATEFORMAT DMY;

        WITH COMUNICADO_DEFERIDO AS (
            SELECT
                CF.Matricula,
                EE.NOME_ENTID AS Participante,
                EE.CPF_CGC AS CPF,
                CAST(CF.DataObito AS DATE) AS Obito,

                MAX(CASE
                    WHEN RH.SituacaoId = 1
                    THEN CAST(RH.DataSituacao AS DATE)
                END) AS Incluido,

                MAX(CASE
                    WHEN RH.SituacaoId = 2
                    THEN CAST(RH.DataSituacao AS DATE)
                END) AS Deferido
            FROM Requerimento.ComunicadoFalecimento CF WITH (NOLOCK)

            LEFT JOIN dbo.CS_FUNCIONARIO FUN
                ON FUN.NUM_MATRICULA = CF.Matricula

            LEFT JOIN dbo.EE_ENTIDADE EE WITH (NOLOCK)
                ON EE.COD_ENTID = FUN.COD_ENTID

            LEFT JOIN Requerimento.HistoricoSituacao RH
                ON RH.RequerimentoId = CF.RequerimentoId

            WHERE 1=1
            AND RH.SituacaoId IN (1, 2)

            GROUP BY
                CF.Matricula,
                EE.NOME_ENTID,
                EE.CPF_CGC,
                CF.DataObito
        )
        SELECT
            Matricula,
            Participante,
            CPF,
            FORMAT(Obito, 'dd/MM/yyyy') AS DATAOBITO,
            FORMAT(Incluido, 'dd/MM/yyyy') AS DATAINCLUSAO,
            FORMAT(Deferido, 'dd/MM/yyyy') AS DATADEFERIMENTO,
            DATEDIFF(dd, Obito, Incluido) AS dias_entre_obito_comunicado,
            DATEDIFF(dd, Incluido, Deferido) AS dias_entre_comunicado_deferimento
        FROM COMUNICADO_DEFERIDO
        WHERE Deferido IS NOT NULL;
    """.strip()

    with engine.connect() as conn:
        df = pd.read_sql(text(query), conn)

    if df.empty:
        return pd.DataFrame(columns=["MATRICULA", "DATAINCLUSAO", "DATADEFERIMENTO"])

    # Força o nome da coluna para maiúsculo para bater com o filtro abaixo
    df.columns = [col.upper() for col in df.columns]

    df["MATRICULA"] = df["MATRICULA"].apply(normalizar_matricula)

    # Se houver mais de um registro por matrícula, mantém o último encontrado.
    df = df.drop_duplicates(subset=["MATRICULA"], keep="last")

    return df[["MATRICULA", "DATAINCLUSAO", "DATADEFERIMENTO"]]


def enriquecer_abas_excel(
    arquivo_excel: Path, df_datas: pd.DataFrame
) -> dict[str, pd.DataFrame]:
    """
    Modificada: Não processa mais os DataFrames diretamente para salvar.
    Agora apenas prepara o mapa de dados. A escrita com formatação será feita no salvamento.
    """
    # Prepara o mapa de datas (continua igual e rápido com Pandas)
    df_datas = df_datas.copy()
    df_datas["DATAINCLUSAO_DT"] = pd.to_datetime(
        df_datas["DATAINCLUSAO"], errors="coerce", dayfirst=True
    )
    df_datas = (
        df_datas.sort_values(["MATRICULA", "DATAINCLUSAO_DT"])
        .drop_duplicates(subset=["MATRICULA"], keep="first")
        .drop(columns=["DATAINCLUSAO_DT"])
    )
    mapa_datas = df_datas.set_index("MATRICULA")[["DATAINCLUSAO", "DATADEFERIMENTO"]]

    # Retornamos o mapa para ser usado diretamente na nova função de salvamento
    return mapa_datas


def salvar_excel_saida(mapa_datas: pd.DataFrame, arquivo_entrada: Path) -> Path:
    out_dir = Path("Arquivos")
    garantir_pasta(str(out_dir))

    base = sanitize_filename(arquivo_entrada.stem)
    ts = time.strftime("%Y%m%d_%H%M%S")
    arquivo_saida = out_dir / f"{base}_com_datas_falecimento_{ts}.xlsx"

    # Carrega o workbook original PRESERVANDO todos os estilos e formatos
    wb = openpyxl.load_workbook(arquivo_entrada)

    for nome_aba in wb.sheetnames:
        ws = wb[nome_aba]

        # Se a aba estiver vazia, pula
        if ws.max_row < 1:
            continue

        # Ler o cabeçalho (primeira linha) para achar a matrícula e verificar se as colunas já existem
        cabecalho = [
            str(ws.cell(row=1, column=col).value) for col in range(1, ws.max_column + 1)
        ]

        # Localiza a coluna de matrícula usando a sua lógica de limpeza
        idx_matricula = None
        candidatos = {"MATRICULA", "NUMMATRICULA", "CDMATRICULA", "NUMEROMATRICULA"}

        for idx, col_name in enumerate(cabecalho, start=1):
            if limpar_nome_coluna(col_name) in candidatos:
                idx_matricula = idx
                break

        if not idx_matricula:
            print(
                f"Aviso: aba '{nome_aba}' sem coluna de matrícula. Ignorando esta aba."
            )
            continue

        # Verifica se as colunas já existem para não duplicar, ou define onde criar
        idx_inclusao = None
        idx_deferimento = None

        for idx, col_name in enumerate(cabecalho, start=1):
            if col_name == "DATAINCLUSAO":
                idx_inclusao = idx
            elif col_name == "DATADEFERIMENTO":
                idx_deferimento = idx

        # Se não existem, adiciona no final do cabeçalho
        if not idx_inclusao:
            ws.cell(row=1, column=ws.max_column + 1, value="DATAINCLUSAO")
            idx_inclusao = ws.max_column
        if not idx_deferimento:
            ws.cell(row=1, column=ws.max_column + 1, value="DATADEFERIMENTO")
            idx_deferimento = ws.max_column

        # Percorre as linhas preenchendo os dados
        for r in range(2, ws.max_row + 1):
            val_matricula = ws.cell(row=r, column=idx_matricula).value

            # Normaliza a matrícula da célula
            matricula_norm = normalizar_matricula(val_matricula)

            data_inc = ""
            data_def = ""

            if matricula_norm in mapa_datas.index:
                data_inc = mapa_datas.loc[matricula_norm, "DATAINCLUSAO"]
                data_def = mapa_datas.loc[matricula_norm, "DATADEFERIMENTO"]

            # Grava apenas os valores nas novas colunas, mantendo o resto intacto
            ws.cell(row=r, column=idx_inclusao, value=data_inc)
            ws.cell(row=r, column=idx_deferimento, value=data_def)

            # Opcional: Forçar o formato de texto (@) para as novas colunas de data não mudarem sozinhos
            ws.cell(row=r, column=idx_inclusao).number_format = "@"
            ws.cell(row=r, column=idx_deferimento).number_format = "@"

    # Salva o novo arquivo
    wb.save(arquivo_saida)
    wb.close()

    return arquivo_saida


def main():
    limpar_console()

    try:
        arquivo_excel = solicitar_arquivo_excel()
    except Exception as e:
        print(f"Erro no arquivo de entrada: {e}")
        return

    try:
        cfg = carregar_cfg()
    except Exception as e:
        print(f"Erro nas variáveis de ambiente: {e}")
        return

    print("Conectando ao banco de dados...")
    try:
        engine = get_engine(cfg)
        with engine.connect() as conn:
            conn.execute(text("SELECT 1"))
        print("Conexão bem-sucedida.")
    except Exception as e:
        print(f"Erro ao conectar ao banco de dados: {e}")
        return

    print("Executando consulta de comunicados de falecimento...")
    try:
        df_datas = buscar_datas_comunicado_falecimento(engine)
        print(f"Registros retornados pela consulta: {len(df_datas)}")
    except Exception as e:
        print(f"Erro ao executar a query: {e}")
        return

    print("Lendo Excel e preparando cruzamento de dados...")
    try:
        # Agora essa função retorna o DataFrame indexado com as datas prontas
        mapa_datas = enriquecer_abas_excel(arquivo_excel, df_datas)
    except Exception as e:
        print(f"Erro ao processar o Excel: {e}")
        return

    print("Gerando arquivo de saída mantendo a formatação original...")
    try:
        # Passamos o mapa de datas e o arquivo original para o openpyxl fazer a cópia formatada
        arquivo_saida = salvar_excel_saida(mapa_datas, arquivo_excel)
        print(f"Arquivo salvo com sucesso: {arquivo_saida}")
    except Exception as e:
        print(f"Erro ao salvar o arquivo Excel: {e}")


if __name__ == "__main__":
    main()
